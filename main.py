#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import json
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QPushButton, QLabel, QLineEdit,
    QFileDialog, QMessageBox, QHeaderView, QToolBar
)
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QKeySequence, QAction


from PyQt6.QtWidgets import QStyledItemDelegate, QLineEdit

class CleanEditDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: black;
            }
        """)
        return editor

def parse_number(s):
    try:
        if isinstance(s, (int, float)):
            return float(s)
        s = str(s).strip().replace(',', '.')
        return float(s) if s else 0.0
    except Exception:
        return 0.0

class DragDropTableWidget(QTableWidget):
    def __init__(self, *args, load_callback=None, **kwargs):
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)
        self.viewport().setAcceptDrops(True)
        self.setDragDropMode(QTableWidget.DragDropMode.DropOnly)
        self.load_callback = load_callback

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if self.load_callback:
                    self.load_callback(file_path)
            event.acceptProposedAction()
        else:
            event.ignore()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Layouts")
        self.current_file = None
        self.resize(1100, 550)

        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout()
        central.setLayout(layout)

        save_over_act = QAction("Сохранить", self)
        save_over_act.setShortcut(QKeySequence("Ctrl+S"))
        save_over_act.triggered.connect(self.save_to_file)

        save_act = QAction("Сохранить как...", self)
        save_act.setShortcut(QKeySequence("Ctrl+Shift+S"))
        save_act.triggered.connect(lambda: self.save_to_file(force_dialog=True))

        load_act = QAction("Загрузить...", self)
        load_act.triggered.connect(self.load_from_file)

        export_act = QAction("Экспорт отчёта (TXT)...", self)
        export_act.triggered.connect(self.export_report)

        exel_act = QAction("Экспорт отчёта (Exel)...", self)
        exel_act.triggered.connect(self.export_excel)

        clear_act = QAction("Очистить таблицу", self)
        clear_act.setShortcut(QKeySequence("Ctrl+L"))
        clear_act.triggered.connect(self.clear_table)

        exit_act = QAction("Выход", self)
        exit_act.triggered.connect(self.close)

        toolbar = QToolBar("Main")
        toolbar.addAction(save_over_act)
        toolbar.addAction(save_act)
        toolbar.addAction(load_act)
        toolbar.addAction(export_act)
        toolbar.addAction(exel_act)
        toolbar.addAction(clear_act)
        toolbar.addAction(exit_act)

        self.addToolBar(toolbar)

        # --- Таблица ---
        # self.table = QTableWidget(0, 4)
        self.table = DragDropTableWidget(0, 4, load_callback=self.load_file_path)
        self.table.setHorizontalHeaderLabels([
            "Название макета",
            "Цена за штуку",
            "Кол-во экземпляров",
            "Общая стоимость"
        ])
        header = self.table.horizontalHeader()
        self.table.setColumnWidth(0, 630)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        for i in range(1, 3):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setItemDelegate(CleanEditDelegate())
        layout.addWidget(self.table)


        # --- Нижняя панель с кнопками и общей суммой ---
        bottom = QHBoxLayout()
        layout.addLayout(bottom)

        btn_add = QPushButton("+")
        btn_add.setToolTip("Добавить запись")
        btn_add.setFixedSize(56, 56)
        btn_add.setStyleSheet("font-size:28px; color:white; background-color:#2ecc71; border-radius:6px;")
        btn_add.clicked.connect(self.add_row)

        btn_remove = QPushButton("−")
        btn_remove.setToolTip("Удалить выбранные записи")
        btn_remove.setFixedSize(56, 56)
        btn_remove.setStyleSheet("font-size:28px; color:white; background-color:#e74c3c; border-radius:6px;")
        btn_remove.clicked.connect(self.remove_selected_rows)

        bottom.addWidget(btn_add)
        bottom.addWidget(btn_remove)
        bottom.addStretch(1)

        lbl_total = QLabel("Общая сумма:")
        lbl_total.setStyleSheet("font-weight:600; color:#555;")
        self.total_edit = QLineEdit("0.00")
        self.total_edit.setReadOnly(True)
        self.total_edit.setFixedWidth(140)
        self.total_edit.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        bottom.addWidget(lbl_total)
        bottom.addWidget(self.total_edit)

        # Сигналы
        self.table.cellChanged.connect(self.on_cell_changed)

        # Стартовая строка
        self.add_row()

    # --------------------------
    # Логика таблицы
    # --------------------------
    def add_row(self, name="", price=0.0, qty=0):
        row = self.table.rowCount()
        self.table.insertRow(row)

        item_name = QTableWidgetItem(name)
        self.table.setItem(row, 0, item_name)

        item_price = QTableWidgetItem("" if price == 0 else f"{price:.2f}")
        item_price.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(row, 1, item_price)

        item_qty = QTableWidgetItem("" if qty == 0 else str(int(qty)))
        item_qty.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.table.setItem(row, 2, item_qty)

        item_total = QTableWidgetItem(f"{price * qty:.2f}")
        item_total.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        item_total.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(row, 3, item_total)

        # self.table.resizeRowToContents(row)
        self.update_overall_sum()

    def remove_selected_rows(self):
        selected = self.table.selectedItems()
        if selected:
            rows = sorted({it.row() for it in selected}, reverse=True)
            for r in rows:
                self.table.removeRow(r)
        elif self.table.rowCount() > 0:
            self.table.removeRow(self.table.rowCount() - 1)
        self.update_overall_sum()

    def on_cell_changed(self, row, col):
        if row < 0 or col not in (1, 2):
            return

        self.table.blockSignals(True)

        price = parse_number(self.table.item(row, 1).text() if self.table.item(row, 1) else "0")
        qty = parse_number(self.table.item(row, 2).text() if self.table.item(row, 2) else "0")
        total = price * qty

        total_item = self.table.item(row, 3)
        if not total_item:
            total_item = QTableWidgetItem()
            total_item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 3, total_item)

        total_item.setText(f"{total:.2f}")
        self.table.blockSignals(False)

        # self.table.resizeRowToContents(row)
        self.update_overall_sum()

    def update_overall_sum(self):
        total_sum = 0.0
        for r in range(self.table.rowCount()):
            price = parse_number(self.table.item(r, 1).text() if self.table.item(r, 1) else "0")
            qty = parse_number(self.table.item(r, 2).text() if self.table.item(r, 2) else "0")
            total_sum += price * qty
        self.total_edit.setText(f"{total_sum:.2f}")

    # --------------------------
    # Сохранение / загрузка / отчёт
    # --------------------------
    def save_to_file(self, force_dialog=False):
        """
        Сохраняет данные таблицы:
        - если путь уже есть (self.current_file) и force_dialog=False — сохраняет поверх;
        - если пути нет, путь указывает на .xml или force_dialog=True — открывает диалог 'Сохранить как...'.
        """
        fn = self.current_file

        # если путь не задан, явно запрошен диалог или расширение .xml — открываем диалог сохранения
        if not fn or fn.lower().endswith(".xml") or force_dialog:
            default_dir = str(Path(fn).parent) if fn else str(Path.home())
            default_name = (
                f"{Path(fn).stem}.json" if fn and Path(fn).suffix.lower() == ".xml" else ""
            )

            fn, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить данные (JSON)",
                str(Path(default_dir) / default_name),
                "JSON Files (*.json)"
            )
            if not fn:
                return
            if not fn.lower().endswith(".json"):
                fn += ".json"

            self.current_file = fn
            self.setWindowTitle(f"Layouts - {Path(fn).name}")

        # сбор данных из таблицы
        data = []
        for r in range(self.table.rowCount()):
            name = self.table.item(r, 0).text() if self.table.item(r, 0) else ""
            price = parse_number(self.table.item(r, 1).text() if self.table.item(r, 1) else "0")
            qty = int(parse_number(self.table.item(r, 2).text() if self.table.item(r, 2) else "0"))
            data.append({"name": name, "price": price, "qty": qty})

        # сохранение
        try:
            with open(self.current_file, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")

    def load_file_path(self, fn):
        try:
            self.table.blockSignals(True)
            self.table.setRowCount(0)

            if fn.lower().endswith(".json"):
                with open(fn, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for entry in data:
                    self.add_row(
                        entry.get("name", ""),
                        entry.get("price", 0.0),
                        entry.get("qty", 0)

                    )

            elif fn.lower().endswith(".xml"):
                tree = ET.parse(fn)
                root = tree.getroot()
                for layout in root.findall("layout"):
                    name = layout.findtext("name", "")
                    qty = int(layout.findtext("count_of_finished", "0"))
                    price = float(layout.findtext("count_of_one_object", "0"))
                    self.add_row(name, price, qty)

            self.table.blockSignals(False)
            self.current_file = fn
            self.setWindowTitle(f"Layouts - {Path(fn).name}")
            self.update_overall_sum()
            self.table.viewport().update()

        except Exception as e:
            self.table.blockSignals(False)
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл:\n{e}")

    def load_from_file(self):
        fn, _ = QFileDialog.getOpenFileName(
            self,
            "Загрузить данные",
            str(Path.home()),
            "JSON Files (*.json);;XML Files (*.xml)"
        )
        if not fn:
            return

        try:
            self.table.blockSignals(True)
            self.table.setRowCount(0)

            if fn.lower().endswith(".json"):
                with open(fn, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for entry in data:
                    self.add_row(
                        entry.get("name", ""),
                        entry.get("price", 0.0),
                        entry.get("qty", 0)
                    )

            elif fn.lower().endswith(".xml"):
                tree = ET.parse(fn)
                root = tree.getroot()
                for layout in root.findall("layout"):
                    name = layout.findtext("name", "")
                    qty = int(layout.findtext("count_of_finished", "0"))
                    price = float(layout.findtext("count_of_one_object", "0"))
                    self.add_row(name, price, qty)

            self.table.blockSignals(False)
            self.current_file = fn
            self.setWindowTitle(f"Layouts - {Path(fn).name}")
            self.update_overall_sum()
            self.table.viewport().update()

        except Exception as e:
            self.table.blockSignals(False)
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл:\n{e}")

    def export_report(self):
        fn, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчёт", str(Path.home()), "Text Files (*.txt)"
        )
        if not fn:
            return

        try:
            lines = []
            lines.append("Отчёт по макетам\n")

            WIDTH_NAME = 50
            WIDTH_PRICE = 12
            WIDTH_QTY = 8
            WIDTH_SUBTOTAL = 15

            header = (
                f"{'Название макета':{WIDTH_NAME}} | "
                f"{'Стоимость':>{WIDTH_PRICE}} | "
                f"{'Кол-во':>{WIDTH_QTY}} | "
                f"{'Итого':>{WIDTH_SUBTOTAL}}"
            )
            lines.append(header)
            lines.append("-" * (WIDTH_NAME + WIDTH_PRICE + WIDTH_QTY + WIDTH_SUBTOTAL + 9))

            total_sum = 0.0
            for r in range(self.table.rowCount()):
                name = (self.table.item(r, 0).text() if self.table.item(r, 0) else "").strip()
                price = parse_number(self.table.item(r, 1).text() if self.table.item(r, 1) else "0")
                qty = int(parse_number(self.table.item(r, 2).text() if self.table.item(r, 2) else "0"))
                subtotal = price * qty
                total_sum += subtotal

                line = f"{name:{WIDTH_NAME}.{WIDTH_NAME}} | {price:>{WIDTH_PRICE}.2f} | {qty:>{WIDTH_QTY}d} | {subtotal:>{WIDTH_SUBTOTAL}.2f}"
                lines.append(line)

            lines.append("-" * (WIDTH_NAME + WIDTH_PRICE + WIDTH_QTY + WIDTH_SUBTOTAL + 9))
            lines.append(
                f"{'Общая сумма:':>{WIDTH_NAME + WIDTH_PRICE + WIDTH_QTY + 3}} {total_sum:>{WIDTH_SUBTOTAL}.2f}")

            # Сохраняем файл
            with open(fn, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить отчёт:\n{e}")

    def clear_table(self):
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы действительно хотите удалить все записи?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self.table.setRowCount(0)
            self.current_file = None
            self.setWindowTitle("Layouts")

    def export_excel(self):
        fn, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчёт Excel", str(Path.home()), "Excel Files (*.xlsx)"
        )
        if not fn:
            return
        if not fn.endswith(".xlsx"):
            fn += ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт макетов"

        # Заголовки
        headers = ["Название макета", "Стоимость за штуку", "Кол-во", "Итого"]
        ws.append(headers)

        # Жирный шрифт для заголовка
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        total_sum = 0.0

        # Данные из таблицы
        for r in range(self.table.rowCount()):
            name = self.table.item(r, 0).text() if self.table.item(r, 0) else ""
            price = float(parse_number(self.table.item(r, 1).text() if self.table.item(r, 1) else "0"))
            qty = int(parse_number(self.table.item(r, 2).text() if self.table.item(r, 2) else "0"))
            subtotal = price * qty
            total_sum += subtotal

            row_data = [name, price, qty, subtotal]
            ws.append(row_data)

            # Выравнивание: текст влево, числа вправо
            ws.cell(row=r + 2, column=1).alignment = Alignment(horizontal="left")
            for c in range(2, 5):
                ws.cell(row=r + 2, column=c).alignment = Alignment(horizontal="right")

        # Итоговая сумма внизу
        ws.append(["", "", "Общая сумма:", total_sum])
        ws.cell(row=self.table.rowCount() + 2, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=self.table.rowCount() + 2, column=4).alignment = Alignment(horizontal="right")
        ws.cell(row=self.table.rowCount() + 2, column=3).font = Font(bold=True)
        ws.cell(row=self.table.rowCount() + 2, column=4).font = Font(bold=True)

        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        wb.save(fn)

def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

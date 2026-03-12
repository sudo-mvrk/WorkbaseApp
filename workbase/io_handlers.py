import json
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def save_json(file_path, data):
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_file(file_path):
    data = []
    fn_lower = file_path.lower()

    if fn_lower.endswith(".json"):
        with open(file_path, "r", encoding="utf-8") as f:
            raw_data = json.load(f)
            for entry in raw_data:
                data.append({
                    "name": entry.get("name", ""),
                    "price": float(entry.get("price", 0.0)),
                    "qty": int(entry.get("qty", 0))
                })
    elif fn_lower.endswith(".xml"):
        tree = ET.parse(file_path)
        root = tree.getroot()
        for layout in root.findall("layout"):
            name = layout.findtext("name", "")
            qty = int(layout.findtext("count_of_finished", "0"))
            price = float(layout.findtext("count_of_one_object", "0"))
            data.append({"name": name, "price": price, "qty": qty})

    return data


def export_txt(file_path, data, total_sum):
    lines = ["Отчёт по макетам\n"]
    WIDTH_NAME, WIDTH_PRICE, WIDTH_QTY, WIDTH_SUBTOTAL = 50, 12, 8, 15

    header = f"{'Название макета':{WIDTH_NAME}} | {'Стоимость':>{WIDTH_PRICE}} | {'Кол-во':>{WIDTH_QTY}} | {'Итого':>{WIDTH_SUBTOTAL}}"
    lines.append(header)
    lines.append("-" * (WIDTH_NAME + WIDTH_PRICE + WIDTH_QTY + WIDTH_SUBTOTAL + 9))

    for row in data:
        name = str(row["name"]).strip()
        price = row["price"]
        qty = row["qty"]
        subtotal = price * qty
        line = f"{name:{WIDTH_NAME}.{WIDTH_NAME}} | {price:>{WIDTH_PRICE}.2f} | {qty:>{WIDTH_QTY}d} | {subtotal:>{WIDTH_SUBTOTAL}.2f}"
        lines.append(line)

    lines.append("-" * (WIDTH_NAME + WIDTH_PRICE + WIDTH_QTY + WIDTH_SUBTOTAL + 9))
    lines.append(f"{'Общая сумма:':>{WIDTH_NAME + WIDTH_PRICE + WIDTH_QTY + 3}} {total_sum:>{WIDTH_SUBTOTAL}.2f}")

    with open(file_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def export_excel(file_path, data, total_sum):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт макетов"

    headers = ["Название макета", "Стоимость за штуку", "Кол-во", "Итого"]
    ws.append(headers)

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(data, start=2):
        name = row["name"]
        price = row["price"]
        qty = row["qty"]
        subtotal = price * qty

        ws.append([name, price, qty, subtotal])
        ws.cell(row=r_idx, column=1).alignment = Alignment(horizontal="left")
        for c in range(2, 5):
            ws.cell(row=r_idx, column=c).alignment = Alignment(horizontal="right")

    last_row = len(data) + 2
    ws.append(["", "", "Общая сумма:", total_sum])
    ws.cell(row=last_row, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=last_row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=last_row, column=3).font = Font(bold=True)
    ws.cell(row=last_row, column=4).font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(file_path)